import os
from os.path import basename
import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv

path_input_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_output_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
files_dir = os.listdir(path_input_files)
path_zip = os.path.join(path_output_files, "archive.zip")


def test_create_archive(clear_dir):
    """Test: create zip archive with files: pdf, xlsx, csv"""
    with zipfile.ZipFile(path_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in files_dir:
            add_file = os.path.join(path_input_files, file)
            zf.write(add_file, basename(add_file))

    files = os.listdir(path_output_files)
    assert len(files) == 1, f"Expected number of archive file(s): {len(files)}; actual number of archive file(s): {1}"


def test_pdf():
    """Test: reading data from extracted pdf file"""
    with zipfile.ZipFile(path_zip) as zf:
        pdf_file = zf.extract("pdf_file.pdf")
        reader = PdfReader(pdf_file)
        page = reader.pages[0]
        text = page.extract_text()
        assert 'Jul 14' in text, f"Expected result: {'Jul 14'}; actual result: {text}"
        os.remove(pdf_file)


def test_xlsx():
    """Test: reading data from extracted xlsx file"""
    with zipfile.ZipFile(path_zip) as zf:
        xf = zf.extract("xlsx_file.xlsx")
        xlsx_file = load_workbook(xf)
        sheet = xlsx_file.active
        # print(sheet.cell(row=21, column=2).value)
        assert sheet.cell(row=21, column=2).value == "Teresa", f"Expected result: {'Teresa'}, " \
                                                               f"actual result: {sheet.cell(row=21, column=2).value}"
        os.remove(xf)


def test_csv():
    """Test: reading data from extracted xlsx csv"""
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("csv_file.csv")
        with open(cf) as csvfile:
            csvfile = csv.reader(csvfile)
            list_csv = []
            for r in csvfile:
                text = "".join(r).replace(";", " ")
                list_csv.append(text)

            assert list_csv[
                       4] == "jenkins46 9346 Mary Jenkins", f"Expected result: {'jenkins46 9346 Mary Jenkin'}, " \
                                                            f"actual result: {list_csv[4]}"
        os.remove(cf)
